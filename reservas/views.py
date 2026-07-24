from urllib import request

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from reservas .models import Sala
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import date
from django.utils.timezone import datetime
import openpyxl
import pandas as pd
from .forms import ReservaForm

from .models import (
    Aluno, RegistroUso, Reserva, Equipamento, Sala,
    PerfilAdm, NotificacaoFichaAusente, Notebook
)
from .forms import ReservaForm

import base64
import json
import io
 
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from PIL import Image
 
from ultralytics import YOLO

MODEL_PATH = settings.BASE_DIR / "modelos" / "best.pt"
 
# Carrega o modelo UMA vez, quando o servidor sobe (evita recarregar a cada requisição)
modelo_yolo = YOLO(str(MODEL_PATH))
 
# Limite de confiança mínima para considerar uma detecção válida (ajuste conforme necessário)
CONFIANCA_MINIMA = 0.4
# helpers

def _professor_requer_aprovacao(user):
    """Retorna True se a conta do professor está marcada como 'requer aprovação'."""
    try:
        return user.perfil_adm.requer_aprovacao
    except PerfilAdm.DoesNotExist:
        return False



# Auth


def home(request):
    return render(request, 'longa.html')


def CriarConta(request):
    if request.method == "POST":
        usuario = request.POST.get('usuario')
        email = request.POST.get('email')
        senha = request.POST.get('senha')
        confirmar = request.POST.get('confirmar_senha')

        dominio_permitido = "@professor.educacao.sp.gov.br"

        if not email.lower().endswith(dominio_permitido):
            messages.error(request, "Erro: Apenas e-mails corporativo SEDUC!")
            return render(request, 'index.html')

        if senha != confirmar:
            messages.error(request, "As senhas não coincidem!")
            return render(request, 'index.html')

        if User.objects.filter(username=usuario).exists():
            messages.error(request, "Este nome de usuário já está em uso.")
            return render(request, 'index.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Este email esta em uso.")
            return render(request, 'index.html')

        user = User.objects.create_user(username=usuario, email=email, password=senha)
        user.save()

        messages.success(request, "Conta criada com sucesso! Faça login.")
        return render(request, 'longa.html')

    return render(request, 'index.html')


def Entrar(request):
    if request.method == "POST":
        usuario_digitado = request.POST.get('usuario')
        senha_digitada = request.POST.get('senha')

        usuario_existe = User.objects.filter(username=usuario_digitado).exists()

        if not usuario_existe:
            messages.error(request, "Usuário não encontrado!")
            return render(request, 'longa.html')

        user = authenticate(request, username=usuario_digitado, password=senha_digitada)

        if user is None:
            messages.error(request, "Senha incorreta!")
            return render(request, 'longa.html')

        login(request, user)
        return redirect('mural')

    return render(request, 'longa.html')


# Mural / Reservas


@login_required
def mural(request):
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    hora_atual = agora.time()

    data_param = request.GET.get('data')
    data_selecionada = data_param if data_param else hoje.strftime('%Y-%m-%d')
    if request.method == "POST":
        data_reserva_str = request.POST.get('data')
        horario_inicio_str = request.POST.get('horario_inicio')
        data_reservae = datetime.strptime(data_reserva_str, '%Y-%m-%d').date()
        hora_inicioe = datetime.strptime(horario_inicio_str, '%H:%M').time()
        equipamento_id = request.POST.get('equipamento')
        try:
            sala_obj = Sala.objects.get(id=request.POST.get('sala'))
        except Sala.DoesNotExist:
            messages.error(request, "Sala não encontrada!")
            return redirect('mural')
        horario_inicio = request.POST.get('horario_inicio')
        horario_fim = request.POST.get('horario_fim')
        data_reserva = request.POST.get('data')
        horario_inicio_obj = datetime.strptime(horario_inicio, '%H:%M').time()
        horario_fim_obj = datetime.strptime(horario_fim, '%H:%M').time()

        professor_reserva = request.user
        professor_id = request.user.id
        

        if data_reservae == hoje and horario_fim_obj < hora_atual:
            messages.error(request, "Este horário já passou e não pode ser reservado!")
            return redirect('/Logar/?data={data_reserva}')

        if request.user.is_staff and request.POST.get('professor'):
            professor_id = request.POST.get('professor')
            try:
                professor_reserva = User.objects.get(id=professor_id)
            except User.DoesNotExist:
                messages.error(request, f"Erro: Professor não encontrado!")
                return redirect('mural')
        
        dupla_reserva = Reserva.objects.filter(
            professor_id = professor_id,
            sala=sala_obj,
            data_uso=data_reservae,
            horario_inicio=horario_inicio_obj,
            horario_fim=horario_fim_obj,

        ).exists()
        ja_reservado = Reserva.objects.filter(
            equipamento_id=equipamento_id,
            data_uso=data_reservae,
            horario_inicio=horario_inicio_obj,
            horario_fim=horario_fim_obj,
            status__in=['confirmada', 'pendente']
        ).exists()

        
        if ja_reservado:
            messages.error(request, "Este carrinho já foi reservado para este horário!")
        elif dupla_reserva:
            messages.error(request, "Você ja tem uma reserva nesse horario e para essa turma!")
            return redirect("mural")
        else:
            try:
                equip_obj = Equipamento.objects.get(id=equipamento_id)

              
                status_reserva = 'confirmada'
                if _professor_requer_aprovacao(professor_reserva):
                    status_reserva = 'pendente'

                Reserva.objects.create(
                    professor=professor_reserva,
                    equipamento=equip_obj,
                    sala=sala_obj,
                    horario_inicio=horario_inicio_obj,
                    horario_fim=horario_fim_obj,
                    data_uso=data_reservae,
                    status=status_reserva,
                )

                if status_reserva == 'pendente':
                    messages.warning(
                        request,
                        f"Reserva enviada! Aguardando aprovação de um administrador."
                    )
                else:
                    messages.success(request, f"Reserva realizada com sucesso para {professor_reserva.username}!")
            except Exception as e:
                messages.error(request, f"Erro ao salvar: {e}")



        return redirect(f"/Logar/?data={data_reserva}")

  
    reservas_hoje = Reserva.objects.filter(
        data_uso=hoje,
        horario_fim__gte=hora_atual,
        status__in=['confirmada', 'pendente']
    ).order_by('horario_inicio')

    equipamentos = Equipamento.objects.all()


    reservas_com_fichas = Reserva.objects.filter(
        registrouso__isnull=False
    ).select_related(
        'professor', 'equipamento'
    ).prefetch_related(
        'registrouso_set__aluno__sala'
    ).distinct().order_by('-data_uso')

    # Reservas pendentes de aprovação (para admins)
    reservas_pendentes = []
    if request.user.is_staff:
        reservas_pendentes = Reserva.objects.filter(
            status='pendente'
        ).select_related('professor', 'equipamento').order_by('data_criacao')

    return render(request, 'mural.html', {
        'reservas': reservas_hoje,
        'equipamentos': equipamentos,
        'hoje': data_selecionada,
        'reservas_com_fichas': reservas_com_fichas,
        'reservas_pendentes': reservas_pendentes,
        'form': ReservaForm(), 
    })


@login_required
@staff_member_required
def aprovar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    reserva.status = 'confirmada'
    reserva.save()
    messages.success(request, f"Reserva de {reserva.professor.username} aprovada!")
    return redirect('mural')


@login_required
@staff_member_required
def recusar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    reserva.status = 'recusada'
    reserva.save()
    messages.warning(request, f"Reserva de {reserva.professor.username} recusada.")
    return redirect('mural')



# Excel / exportar


@login_required
def exportar_reservas_excel(request):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
 
    reservas = Reserva.objects.all().select_related(
        'professor', 'equipamento', 'sala'
    ).order_by('data_uso', 'horario_inicio')
 
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # remove aba padrão vazia
 
    # Estilos
    cor_header     = PatternFill("solid", fgColor="1B3A5C")
    cor_confirmada = PatternFill("solid", fgColor="E6F4EA")
    cor_pendente   = PatternFill("solid", fgColor="FFF8E1")
    cor_recusada   = PatternFill("solid", fgColor="FDECEA")
    fonte_header   = Font(bold=True, color="FFFFFF", size=11)
    fonte_normal   = Font(size=10)
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    centralizado = Alignment(horizontal='center', vertical='center')
 
    colunas  = ['Data', 'Professor', 'Equipamento', 'Sala', 'Início', 'Fim', 'Status']
    larguras = [14,     24,          18,             10,     10,       10,    16]
 
    # Agrupa por mês
    por_mes = defaultdict(list)
    for r in reservas:
        chave = r.data_uso.strftime('%m-%Y')
        por_mes[chave].append(r)
 
    if not por_mes:
        ws = workbook.create_sheet("Sem dados")
        ws.append(["Nenhuma reserva encontrada."])
    else:
        for mes, lista in sorted(por_mes.items()):
            nome_aba = f"Reservas {mes}"[:31]
            ws = workbook.create_sheet(nome_aba)
 
            # Cabeçalho
            ws.append(colunas)
            for col_num, (_, larg) in enumerate(zip(colunas, larguras), 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = cor_header
                cell.font = fonte_header
                cell.alignment = centralizado
                cell.border = borda
                ws.column_dimensions[get_column_letter(col_num)].width = larg
            ws.row_dimensions[1].height = 22
 
            # Linhas de dados
            for row_num, r in enumerate(lista, 2):
                status = r.status
                cor_linha = (cor_confirmada if status == 'confirmada'
                             else cor_pendente if status == 'pendente'
                             else cor_recusada)
                status_label = {
                    'confirmada': '✓ Confirmada',
                    'pendente':   '⏳ Pendente',
                    'recusada':   '✗ Recusada',
                }.get(status, status)
 
                ws.append([
                    r.data_uso.strftime('%d/%m/%Y'),
                    r.professor.get_full_name() or r.professor.username,
                    r.equipamento.nome,
                    r.sala.nome,
                    r.horario_inicio.strftime('%H:%M'),
                    r.horario_fim.strftime('%H:%M'),
                    status_label,
                ])
                for col_num in range(1, len(colunas) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = cor_linha
                    cell.font = fonte_normal
                    cell.border = borda
                    cell.alignment = Alignment(vertical='center')
                ws.row_dimensions[row_num].height = 18
 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reservas.xlsx'
    workbook.save(response)
    return response


@login_required
@staff_member_required
def painel_reservas_dia(request):
    """Painel de cards das reservas — filtrável por data."""
    data_param = request.GET.get('data')
    try:
        data_sel = datetime.strptime(data_param, '%Y-%m-%d').date() if data_param else date.today()
    except ValueError:
        data_sel = date.today()

    reservas = Reserva.objects.filter(
        data_uso=data_sel
    ).select_related('professor', 'equipamento', 'sala').order_by('horario_inicio')

    return render(request, 'painel_reservas.html', {
        'reservas': reservas,
        'data_atual': data_sel.strftime('%Y-%m-%d'),
        'data_exibicao': data_sel.strftime('%d/%m/%Y'),
    })


@login_required
def excluir_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)

    if request.user == reserva.professor or request.user.is_staff:
        reserva.delete()
        messages.success(request, "Reserva excluída com sucesso!")
        

    return redirect('mural')


def listar_disponiveis(request):
    data_sel = request.GET.get('data')
    horario_inicio_sel = request.GET.get('horario_inicio')
    horario_fim_sel = request.GET.get('horario_fim')

    try:
        data_sel_obj = datetime.strptime(data_sel, '%Y-%m-%d').date()
        horario_inicio_obj = datetime.strptime(horario_inicio_sel, '%H:%M').time()
        horario_fim_obj = datetime.strptime(horario_fim_sel, '%H:%M').time()

        ocupados = Reserva.objects.filter(
            data_uso=data_sel_obj,
            horario_inicio=horario_inicio_obj,
            horario_fim=horario_fim_obj,
            status__in=['confirmada', 'pendente']
        ).values_list('equipamento_id', flat=True)

        disponiveis = Equipamento.objects.exclude(id__in=ocupados)
        data = [{'id': e.id, 'nome': e.nome, 'quantidade': e.quantidade} for e in disponiveis]
        return JsonResponse({'equipamentos': data})

    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=500)


def carregar_mural(request):
    data_sel = request.GET.get('data')

    if not data_sel:
        data_sel = date.today().strftime('%Y-%m-%d')

    data_sel_obj = datetime.strptime(data_sel, '%Y-%m-%d').date() if isinstance(data_sel, str) else data_sel

    agora_local = timezone.localtime(timezone.now())
    hora_atual = agora_local.time()
    hoje = agora_local.date()

    if data_sel_obj == hoje:
        reservas = Reserva.objects.filter(
            data_uso=data_sel_obj,
            horario_fim__gte=hora_atual,
            status__in=['confirmada', 'pendente']
        ).order_by('horario_inicio')
    else:
        reservas = Reserva.objects.filter(
            data_uso=data_sel_obj,
            status__in=['confirmada', 'pendente']
        ).order_by('horario_inicio')

    return render(request, 'partials/lista_reservas.html', {
        'reservas': reservas,
        'user': request.user,
    })


def carregar_mural_publico(request):
    data_sel = request.GET.get('data')

    if not data_sel:
        data_sel = date.today().strftime('%Y-%m-%d')

    data_sel_obj = datetime.strptime(data_sel, '%Y-%m-%d').date() if isinstance(data_sel, str) else data_sel

    agora_local = timezone.localtime(timezone.now())
    hora_atual = agora_local.time()
    hoje = agora_local.date()

    if data_sel_obj == hoje:
        reservas = Reserva.objects.filter(
            data_uso=data_sel_obj,
            horario_fim__gte=hora_atual,
            status='confirmada'
        ).order_by('horario_inicio')
    else:
        reservas = Reserva.objects.filter(
            data_uso=data_sel_obj,
            status='confirmada'
        ).order_by('horario_inicio')

    return render(request, 'partials/lista_reservas_consulta.html', {'reservas': reservas,})


def mural_principal(request):
    hoje = date.today()
    hora_atual = timezone.localtime(timezone.now()).time()
    data_sel = date.today().strftime('%Y-%m-%d')
    data_sel_obj = datetime.strptime(data_sel, '%Y-%m-%d').date()

    reservas_hoje = Reserva.objects.filter(
        data_uso=data_sel_obj,
        horario_fim__gte=hora_atual,
        status='confirmada'
    ).order_by('horario_inicio')

    return render(request, 'mural_consulta.html', {
        'hoje': hoje.strftime('%d/%m/%Y'),
        'reservas': reservas_hoje
    })


@login_required
def atualizar_quantidade(request):
    if not request.user.is_staff:
        messages.error(request, "Sem permissão.")
        return redirect('mural')

    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        quantidade = request.POST.get('quantidade')
        try:
            equip = Equipamento.objects.get(id=equipamento_id)
            equip.quantidade = int(quantidade)
            equip.save()
            messages.success(request, f"Quantidade de '{equip.nome}' atualizada!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('mural')


def importar_de_excel(caminho_arquivo):
    df = pd.read_excel(caminho_arquivo)
    for index, row in df.iterrows():
        Sala, _ = Sala.objects.get_or_create(nome=row['sala'])
        Aluno.objects.get_or_create(nome=row['nome'], sala = Sala)


# Tablet / Fichas


def view_tablet(request, equipamento_id):
    agora = timezone.localtime()

    reserva = Reserva.objects.filter(
        equipamento_id=equipamento_id,
        status='confirmada',
        data_uso=agora.date(),
        horario_inicio__lte=agora.time(),
        horario_fim__gte=agora.time(),
    ).first()

    if not reserva:
        return HttpResponse("Nenhuma reserva ativa para este carrinho agora.")

    sala = reserva.sala
    alunos = Aluno.objects.filter(sala=sala)

    if request.method == "POST":
        numeros_usados = {}
        erros = False
        carrinho_id = reserva.equipamento.id  

        for aluno in alunos:
            numero_str = request.POST.get(f'aluno_{aluno.id}')
            if numero_str:
                num = int(numero_str)

                #regra de validação para mandar as fichas
                if carrinho_id == 1 and num > 30:
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} , mais so tem notebooks de 1 a 30.")
                    erros = True
                elif carrinho_id == 10 and (num < 30 or num > 60):
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} mais so tem notebooks de 30 a 60..")
                    erros = True
                elif carrinho_id == 11 and (num < 60 or num > 90):
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} mais so tem notebooks de 60 a 90.")
                    erros = True
                elif carrinho_id == 12 and (num < 90 or num > 120):
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} mais so tem notebooks de 90 a 120.")
                    erros = True
                elif carrinho_id == 13 and (num < 120 or num > 150):
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} mais so tem notebooks de 120 a 150.")
                    erros = True
                elif carrinho_id == 14 and (num < 1 or num > 20):
                    messages.error(request, f"Erro: {aluno.nome} colocou {num} mais so tem notebooks de 120 a 150.")
                    erros = True
            
                if num in numeros_usados:
                    messages.error(
                        request,
                        f"Erro: {aluno.nome} e {numeros_usados[num]} colocaram o mesmo número ({num})!"
                    )
                    erros = True
                else:
                    numeros_usados[num] = aluno.nome
        

      
        if erros:
            return render(request, 'tablet_checkin.html', {'reserva': reserva, 'alunos': alunos, 'dados_anteriores': request.POST})

      
        for aluno in alunos:
            numero = request.POST.get(f'aluno_{aluno.id}')
            if numero:
                RegistroUso.objects.update_or_create(
                    reserva=reserva, aluno=aluno,
                    defaults={'numero_notebook': numero}
                )

        return render(request, 'enviado.html', {
            'id': equipamento_id,
            'reserva_id': reserva.id,
            'professor_id': reserva.professor_id,
            'carrinho_id': carrinho_id,
            'sala': sala.id,
        })

    return render(request, 'tablet_checkin.html', {'reserva': reserva, 'alunos': alunos})


def status_tablet(request, equipamento_id):
    agora = timezone.localtime()

    nova_reserva = Reserva.objects.filter(
        equipamento_id=equipamento_id,
        status__in=['confirmada','pedente'],
        data_uso=agora.date(),
        horario_inicio__lte=agora.time(),
        horario_fim__gte=agora.time(),
    ).exclude(id=request.GET.get('reserva_atual')).exists()

    return JsonResponse({'nova_reserva': nova_reserva})

@staff_member_required
def exportar_todas_fichas(request):
    reservas = Reserva.objects.filter(
        registrouso__isnull=False
    ).distinct().order_by('data_uso')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="todas_fichas.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        alguma_aba = False
        for reserva in reservas:
            registros = RegistroUso.objects.filter(
                reserva=reserva
            ).select_related('aluno__sala', 'reserva__professor')  

            if not registros.exists():
                continue

            data = [{
                'Professor': reserva.professor.get_full_name() or reserva.professor.username,
                'Sala': reg.aluno.sala.nome,       
                'Aluno': reg.aluno.nome,
                'Nº Notebook': reg.numero_notebook,
                'Data': reserva.data_uso,
            } for reg in registros]

            df = pd.DataFrame(data)
            sala = registros.first().aluno.sala.nome  
            aba = f"{sala} {reserva.data_uso.strftime('%d-%m')}"[:31]
            df.to_excel(writer, index=False, sheet_name=aba)
            alguma_aba = True

        if not alguma_aba:
            pd.DataFrame([{'Aviso': 'Nenhuma ficha preenchida ainda.'}]).to_excel(
                writer, index=False, sheet_name='Sem dados'
            )

    return response

@staff_member_required
def exportar_ficha_excel(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    registros = RegistroUso.objects.filter(
        reserva=reserva
    ).select_related('aluno__sala', 'reserva__professor')

    data = [{
        'Professor': reserva.professor.get_full_name() or reserva.professor.username,
        'Sala': reserva.sala.nome,
        'Aluno': reg.aluno.nome,
        'Nº Notebook': reg.numero_notebook,
        'Data': reserva.data_uso,
    } for reg in registros]

    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ficha_{reserva_id}.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ficha')

    return response


@staff_member_required
def painel_fichas(request):
    data_param = request.GET.get('data')
    data_sel = None
    if data_param:
        try:
            data_sel = datetime.strptime(data_param, '%Y-%m-%d').date()
        except ValueError:
            data_sel = None

    reservas = Reserva.objects.filter(
        registrouso__isnull=False
    ).select_related(
        'professor', 'equipamento'
    ).prefetch_related(
        'registrouso_set__aluno__sala'
    ).distinct().order_by('-data_uso')

    if data_sel:
        reservas = reservas.filter(data_uso=data_sel)

    return render(request, 'painelestudante.html', {
        'reservas': reservas,
        'data_selecionada': data_param or '',
    })


@staff_member_required
def ficha_detalhe_json(request, reserva_id):
    registros = RegistroUso.objects.filter(
        reserva_id=reserva_id
    ).select_related('aluno__sala', 'reserva__professor')

    reserva = registros.first().reserva if registros.exists() else None

    if not reserva:
        return JsonResponse({'erro': 'Reserva não encontrada'}, status=404)

    data = {
        'professor': reserva.professor.get_full_name() or reserva.professor.username,
        'data': reserva.data_uso.strftime('%d/%m/%Y'),
        'horario': f"{reserva.horario_inicio.strftime('%H:%M')} - {reserva.horario_fim.strftime('%H:%M')}",
        'sala': reserva.sala.nome,
        'equipamento': reserva.equipamento.nome,
        'registros': [
            {
                'aluno': reg.aluno.nome,
                'turma': reg.aluno.sala.nome,
                'notebook': reg.numero_notebook or '-',
            }
            for reg in registros.order_by('aluno__nome')
        ]
    }
    return JsonResponse(data)


# Notificações de fichas ausentes (chamada via AJAX do frontend admin)


@staff_member_required
def verificar_fichas_ausentes(request):
    """
    Verifica reservas cujo horário já encerrou e não têm ficha preenchida.
    Retorna JSON com lista de pendências para exibir como notificação no admin.
    """
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    hora_atual = agora.time()


    reservas_sem_ficha = Reserva.objects.filter(
        data_uso=hoje,
        horario_fim__lt=hora_atual,
        status='confirmada',
        registrouso__isnull=True
    ).select_related('professor', 'equipamento').distinct()

    pendencias = [
        {
            'id': r.id,
            'professor': r.professor.get_full_name() or r.professor.username,
            'sala': r.sala.nome,
            'horario_fim': r.horario_fim.strftime('%H:%M'),
            'equipamento': r.equipamento.nome,
        }
        for r in reservas_sem_ficha
    ]

    return JsonResponse({'pendencias': pendencias})

@login_required
@staff_member_required
def verificar_carrinho(request):
    """
    Permite que um admin escolha um carrinho e digite os números de notebook
    encontrados dentro dele. Compara com a faixa esperada daquele carrinho
    (numero_inicial..numero_final) e aponta:
      - números corretos (estão na faixa do carrinho)
      - faltando (esperados, mas não digitados)
      - fora de lugar (digitados, mas pertencem à faixa de outro carrinho)
      - não reconhecidos (digitados, mas não pertencem a nenhuma faixa cadastrada)
      - duplicados (digitados mais de uma vez)
    """
    import re

    equipamentos = Equipamento.objects.all().order_by('nome')
    resultado = None
    carrinho_selecionado = None

    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numeros_texto = request.POST.get('numeros', '')
        carrinho_selecionado = get_object_or_404(Equipamento, id=equipamento_id)

        numeros_informados = [int(n) for n in re.findall(r'\d+', numeros_texto)]

        contagem = {}
        for n in numeros_informados:
            contagem[n] = contagem.get(n, 0) + 1
        duplicados = sorted([n for n, c in contagem.items() if c > 1])

        numeros_unicos = set(numeros_informados)
        faixa_esperada = carrinho_selecionado.faixa_numeros()

        faltando = sorted(faixa_esperada - numeros_unicos)

        outros_equipamentos = list(equipamentos.exclude(id=carrinho_selecionado.id))

        corretos = []
        fora_de_lugar = []
        nao_reconhecidos = []

        for n in sorted(numeros_unicos):
            if n in faixa_esperada:
                corretos.append(n)
                continue
            dono = next((o for o in outros_equipamentos if n in o.faixa_numeros()), None)
            if dono:
                fora_de_lugar.append({'numero': n, 'pertence_a': dono.nome})
            else:
                nao_reconhecidos.append(n)

        resultado = {
            'corretos': corretos,
            'faltando': faltando,
            'fora_de_lugar': fora_de_lugar,
            'nao_reconhecidos': nao_reconhecidos,
            'duplicados': duplicados,
            'total_esperado': len(faixa_esperada),
            'total_informado': len(numeros_unicos),
        }

    return render(request, 'verificar_carrinho.html', {
        'equipamentos': equipamentos,
        'resultado': resultado,
        'carrinho_selecionado': carrinho_selecionado,
    })

@login_required
@staff_member_required
def atualizar_faixa_numeracao(request):
    """Atualiza numero_inicial/numero_final de um carrinho a partir da página
    de Verificar Carrinhos, e volta para ela (não para o mural)."""
    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numero_inicial = request.POST.get('numero_inicial')
        numero_final = request.POST.get('numero_final')
        try:
            equip = Equipamento.objects.get(id=equipamento_id)
            equip.numero_inicial = int(numero_inicial) if numero_inicial else None
            equip.numero_final = int(numero_final) if numero_final else None
            equip.save()
            messages.success(request, f"Faixa de numeração de '{equip.nome}' atualizada!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('verificar_carrinho')


@login_required
@staff_member_required
def alternar_status_notebook(request):
    """Marca/desmarca um número de notebook específico como quebrado, dentro de um carrinho."""
    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numero = request.POST.get('numero')
        try:
            equip = Equipamento.objects.get(id=equipamento_id)
            numero_int = int(numero)

            notebook, created = Notebook.objects.get_or_create(
                equipamento=equip,
                numero=numero_int,
                defaults={'ativo': False},
            )
            if not created:
                notebook.ativo = not notebook.ativo
                notebook.save()

            status = "ativo" if notebook.ativo else "quebrado"
            messages.success(request, f"Notebook {numero_int} do '{equip.nome}' marcado como {status}.")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('verificar_carrinho')


@require_POST
@csrf_protect
def analisar_foto(request):
    """
    Recebe um POST em JSON no formato:
        { "imagem": "data:image/jpeg;base64,...", "tipo_carrinho": "notebook" }
 
    Retorna:
        { "contagens": {"notebook": 5, "tablet": 0}, "total": 5 }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"erro": "JSON inválido."}, status=400)
 
    imagem_base64 = body.get("imagem")
    tipo_carrinho = body.get("tipo_carrinho", "")
 
    if not imagem_base64:
        return JsonResponse({"erro": "Nenhuma imagem enviada."}, status=400)
 
    # A string vem como "data:image/jpeg;base64,XXXXX" — precisamos remover o cabeçalho
    try:
        if "," in imagem_base64:
            imagem_base64 = imagem_base64.split(",", 1)[1]
        imagem_bytes = base64.b64decode(imagem_base64)
        imagem = Image.open(io.BytesIO(imagem_bytes)).convert("RGB")
    except Exception:
        return JsonResponse({"erro": "Não foi possível processar a imagem."}, status=400)
 
    # Roda a inferência do YOLO na imagem recebida
    resultados = modelo_yolo.predict(
        source=imagem,
        conf=CONFIANCA_MINIMA,
        verbose=False,
    )
 
    # resultados é uma lista (1 imagem = 1 resultado). Pegamos o primeiro.
    deteccoes = resultados[0]
 
    # Conta quantas detecções de cada classe apareceram
    contagens = {}
    for box in deteccoes.boxes:
        classe_id = int(box.cls[0])
        nome_classe = modelo_yolo.names[classe_id]  # ex: "notebook" ou "tablet"
        contagens[nome_classe] = contagens.get(nome_classe, 0) + 1
 
    total = sum(contagens.values())
 
    return JsonResponse({
        "contagens": contagens,
        "total": total,
        "tipo_carrinho": tipo_carrinho,
    })

@login_required
def camera_contagem(request):
    return render(request, 'camera.html')
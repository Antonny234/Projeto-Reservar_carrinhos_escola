"""
Comando de gerenciamento Django: importar_alunos

Lê a planilha de chamada (uma aba por turma) e cria/atualiza os registros de
Aluno, ligando cada um à Sala correspondente (a Sala já deve existir no banco,
com o nome batendo com o nome da aba, ex: "7ºA", "1ºB", etc).

Uso:
    python manage.py importar_alunos caminho/para/Chamada_2026__NOVA.xlsx
    python manage.py importar_alunos caminho/para/Chamada_2026__NOVA.xlsx --dry-run

Coloque este arquivo em:
    reservas/management/commands/importar_alunos.py

(crie os arquivos __init__.py vazios em reservas/management/ e
reservas/management/commands/ se ainda não existirem — o Django precisa
deles para reconhecer o pacote).
"""
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError

from reservas.models import Aluno, Sala

try:
    import openpyxl
except ImportError:
    openpyxl = None


def normalizar(texto):
    """Remove acentos, espaços e deixa maiúsculo, pra comparar 'nome da aba'
    com 'nome da sala' mesmo se um tiver 'º' e outro 'ª', espaços extras etc."""
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Z0-9]", "", texto)
    return texto


class Command(BaseCommand):
    help = "Importa alunos da planilha de chamada, ligando cada um à Sala correspondente à sua turma."

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho para o arquivo .xlsx")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Só mostra o que seria feito, sem gravar nada no banco.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl não está instalado. Rode: pip install openpyxl")

        caminho = options["arquivo"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Arquivo não encontrado: {caminho}")

        # Monta um dicionário {nome_normalizado_da_sala: objeto Sala}
        salas_por_nome_norm = {normalizar(s.nome): s for s in Sala.objects.all()}

        if not salas_por_nome_norm:
            raise CommandError(
                "Não há nenhuma Sala cadastrada no banco. Cadastre as salas primeiro."
            )

        total_criados = 0
        total_existentes = 0
        abas_sem_sala = []

        for nome_aba in wb.sheetnames:
            sheet = wb[nome_aba]

            sala = salas_por_nome_norm.get(normalizar(nome_aba))
            if sala is None:
                abas_sem_sala.append(nome_aba)
                self.stdout.write(
                    self.style.WARNING(
                        f"[PULADA] Aba '{nome_aba}': nenhuma Sala com esse nome foi encontrada no banco."
                    )
                )
                continue

            nomes_alunos = self._extrair_nomes(sheet)

            if not nomes_alunos:
                self.stdout.write(
                    self.style.WARNING(f"[VAZIA] Aba '{nome_aba}': nenhum nome de aluno encontrado.")
                )
                continue

            self.stdout.write(f"Aba '{nome_aba}' -> Sala '{sala.nome}': {len(nomes_alunos)} aluno(s)")

            for nome in nomes_alunos:
                if dry_run:
                    self.stdout.write(f"    (dry-run) {nome}")
                    continue

                aluno, criado = Aluno.objects.get_or_create(nome=nome, sala=sala)
                if criado:
                    total_criados += 1
                else:
                    total_existentes += 1

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.SUCCESS("Simulação concluída (--dry-run), nada foi gravado."))
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Concluído. {total_criados} aluno(s) criado(s), "
                    f"{total_existentes} já existiam."
                )
            )

        if abas_sem_sala:
            self.stdout.write(
                self.style.WARNING(
                    "Abas sem Sala correspondente no banco (verifique o nome da Sala): "
                    + ", ".join(abas_sem_sala)
                )
            )

    def _extrair_nomes(self, sheet):
        """Encontra a linha de cabeçalho ('NOME DO ALUNO') e lê os nomes na
        coluna B a partir dali, ignorando linhas em branco."""
        linha_cabecalho = None
        for row in sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and normalizar(cell.value) == normalizar("NOME DO ALUNO"):
                    linha_cabecalho = cell.row
                    coluna_nome = cell.column
                    break
            if linha_cabecalho:
                break

        if linha_cabecalho is None:
            # Fallback: assume o layout padrão (cabeçalho na linha 5, nomes na coluna B)
            linha_cabecalho = 5
            coluna_nome = 2

        nomes = []
        vistos = set()
        for row in sheet.iter_rows(min_row=linha_cabecalho + 1, max_col=coluna_nome):
            valor = row[coluna_nome - 1].value
            if valor is None:
                continue
            nome = str(valor).strip()
            if not nome:
                continue
            chave = normalizar(nome)
            if chave in vistos:
                continue
            vistos.add(chave)
            nomes.append(nome)

        return nomes
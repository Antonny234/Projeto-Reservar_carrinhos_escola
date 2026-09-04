from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from datetime import datetime, date, timedelta
from django.utils import timezone
import openpyxl
import pandas as pd
import re
from collections import defaultdict
from django.contrib.auth.models import User
from django.db.models import Count
from django.db.models.functions import ExtractWeekDay


#helpers
from .helpers import enviar_telegram

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import (
    Reserva, RegistroUso, Equipamento, Sala, Notebook,NotificacaoFichaAusente
)
from ..forms import ReservaFixaForm
import uuid

@staff_member_required
def exportar_todas_fichas(request):
    reservas = Reserva.objects.filter(
        registrouso__isnull=False
    ).distinct().order_by('data_uso')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="todas_fichas.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        alguma_aba = False
        for reserva in reservas:
            registros = RegistroUso.objects.filter(
                reserva=reserva
            ).select_related('aluno__sala', 'reserva__professor')  

            if not registros.exists():
                continue

            data = [{
                'Professor': reserva.professor.get_full_name() or reserva.professor.username,
                'Sala': reg.aluno.sala.nome,       
                'Aluno': reg.aluno.nome,
                'Nº Notebook': reg.numero_notebook,
                'Data': reserva.data_uso,
            } for reg in registros]

            df = pd.DataFrame(data)
            sala = registros.first().aluno.sala.nome  
            aba = f"{sala} {reserva.data_uso.strftime('%d-%m')}"[:31]
            df.to_excel(writer, index=False, sheet_name=aba)
            alguma_aba = True

        if not alguma_aba:
            pd.DataFrame([{'Aviso': 'Nenhuma ficha preenchida ainda.'}]).to_excel(
                writer, index=False, sheet_name='Sem dados'
            )

    return response

@staff_member_required
def exportar_ficha_excel(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    registros = RegistroUso.objects.filter(
        reserva=reserva
    ).select_related('aluno__sala', 'reserva__professor')

    data = [{
        'Professor': reserva.professor.get_full_name() or reserva.professor.username,
        'Sala': reserva.sala.nome,
        'Aluno': reg.aluno.nome,
        'Nº Notebook': reg.numero_notebook,
        'Data': reserva.data_uso,
    } for reg in registros]

    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ficha_{reserva_id}.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ficha')

    return response

@login_required
@staff_member_required
def exportar_reservas_excel(request):
 
    reservas = Reserva.objects.all().select_related(
        'professor', 'equipamento', 'sala'
    ).order_by('data_uso', 'horario_inicio')
 
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # remove aba padrão vazia
 
    # Estilos
    cor_header     = PatternFill("solid", fgColor="1B3A5C")
    cor_confirmada = PatternFill("solid", fgColor="E6F4EA")
    cor_pendente   = PatternFill("solid", fgColor="FFF8E1")
    cor_recusada   = PatternFill("solid", fgColor="FDECEA")
    fonte_header   = Font(bold=True, color="FFFFFF", size=11)
    fonte_normal   = Font(size=10)
    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    centralizado = Alignment(horizontal='center', vertical='center')
 
    colunas  = ['Data', 'Professor', 'Equipamento', 'Sala', 'Início', 'Fim', 'Status']
    larguras = [14,     24,          18,             10,     10,       10,    16]
 
    # Agrupa por mês
    por_mes = defaultdict(list)
    for r in reservas:
        chave = r.data_uso.strftime('%m-%Y')
        por_mes[chave].append(r)
 
    if not por_mes:
        ws = workbook.create_sheet("Sem dados")
        ws.append(["Nenhuma reserva encontrada."])
    else:
        for mes, lista in sorted(por_mes.items()):
            nome_aba = f"Reservas {mes}"[:31]
            ws = workbook.create_sheet(nome_aba)
 
            # Cabeçalho
            ws.append(colunas)
            for col_num, (_, larg) in enumerate(zip(colunas, larguras), 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = cor_header
                cell.font = fonte_header
                cell.alignment = centralizado
                cell.border = borda
                ws.column_dimensions[get_column_letter(col_num)].width = larg
            ws.row_dimensions[1].height = 22
 
            # Linhas de dados
            for row_num, r in enumerate(lista, 2):
                status = r.status
                cor_linha = (cor_confirmada if status == 'confirmada'
                             else cor_pendente if status == 'pendente'
                             else cor_recusada)
                status_label = {
                    'confirmada': '✓ Confirmada',
                    'pendente':   '⏳ Pendente',
                    'recusada':   '✗ Recusada',
                }.get(status, status)
 
                ws.append([
                    r.data_uso.strftime('%d/%m/%Y'),
                    r.professor.get_full_name() or r.professor.username,
                    r.equipamento.nome,
                    r.sala.nome,
                    r.horario_inicio.strftime('%H:%M'),
                    r.horario_fim.strftime('%H:%M'),
                    status_label,
                ])
                for col_num in range(1, len(colunas) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = cor_linha
                    cell.font = fonte_normal
                    cell.border = borda
                    cell.alignment = Alignment(vertical='center')
                ws.row_dimensions[row_num].height = 18
 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reservas.xlsx'
    workbook.save(response)
    return response

@staff_member_required
def painel_fichas(request):
    data_param = request.GET.get('data')
    data_sel = None
    if data_param:
        try:
            data_sel = datetime.strptime(data_param, '%Y-%m-%d').date()
        except ValueError:
            data_sel = None

    reservas = Reserva.objects.filter(
        registrouso__isnull=False
    ).select_related(
        'professor', 'equipamento'
    ).prefetch_related(
        'registrouso_set__aluno__sala'
    ).distinct().order_by('-data_uso')

    if data_sel:
        reservas = reservas.filter(data_uso=data_sel)

    return render(request, 'painelestudante.html', {
        'reservas': reservas,
        'data_selecionada': data_param or '',
    })


@staff_member_required
def ficha_detalhe_json(request, reserva_id):
    registros = RegistroUso.objects.filter(
        reserva_id=reserva_id
    ).select_related('aluno__sala', 'reserva__professor')

    reserva = registros.first().reserva if registros.exists() else None

    if not reserva:
        return JsonResponse({'erro': 'Reserva não encontrada'}, status=404)

    data = {
        'professor': reserva.professor.get_full_name() or reserva.professor.username,
        'data': reserva.data_uso.strftime('%d/%m/%Y'),
        'horario': f"{reserva.horario_inicio.strftime('%H:%M')} - {reserva.horario_fim.strftime('%H:%M')}",
        'sala': reserva.sala.nome,
        'equipamento': reserva.equipamento.nome,
        'registros': [
            {
                'aluno': reg.aluno.nome,
                'turma': reg.aluno.sala.nome,
                'notebook': reg.numero_notebook or '-',
            }
            for reg in registros.order_by('aluno__nome')
        ]
    }
    return JsonResponse(data)


# Notificações de fichas ausentes (chamada via AJAX do frontend admin)


@staff_member_required
def verificar_fichas_ausentes(request):
    agora = timezone.localtime(timezone.now())
    hoje = agora.date()
    hora_atual = agora.time()

    reservas_sem_ficha = Reserva.objects.filter(
        data_uso=hoje,
        horario_fim__lt=hora_atual,
        status='confirmada',
        registrouso__isnull=True,
        notificacaofichaausente__isnull=True,   # ainda não tem notificação registrada
    ).select_related('professor', 'equipamento', 'sala').distinct()

    pendencias = []
    for r in reservas_sem_ficha:
        pendencias.append({
            'id': r.id,
            'professor': r.professor.get_full_name() or r.professor.username,
            'sala': r.sala.nome,
            'horario_fim': r.horario_fim.strftime('%H:%M'),
            'equipamento': r.equipamento.nome,
        })

        enviar_telegram(
            f"⚠️ <b>Ficha não preenchida</b>\n"
            f"Professor: {r.professor.get_full_name() or r.professor.username}\n"
            f"Sala: {r.sala.nome}\n"
            f"Carrinho: {r.equipamento.nome}\n"
            f"Encerrou às: {r.horario_fim.strftime('%H:%M')}"
        )

        NotificacaoFichaAusente.objects.create(reserva=r)

    return JsonResponse({'pendencias': pendencias})

@login_required
@staff_member_required
def notebooks_quebrados(request):
    """Página admin que lista todos os notebooks denunciados como quebrados,
    agrupados por carrinho, com ação para marcá-los como consertados."""
    # Agrupa os notebooks inativos por carrinho
    grupos = defaultdict(list)
    quebrados = Notebook.objects.filter(
        ativo=False
    ).select_related('equipamento').order_by('equipamento__nome', 'numero')

    for nb in quebrados:
        grupos[nb.equipamento].append(nb)

    total_quebrados = quebrados.count()

    return render(request, 'notebooks_quebrados.html', {
        'grupos': dict(grupos),
        'total_quebrados': total_quebrados,
    })


@login_required
@staff_member_required
def reativar_notebook(request):
    """Marca um notebook denunciado como quebrado de volta para 'ativo' (consertado)."""
    if request.method != "POST":
        messages.error(request, "Método inválido.")
        return redirect('notebooks_quebrados')

    notebook_id = request.POST.get('notebook_id')
    try:
        notebook = get_object_or_404(Notebook, id=notebook_id)
        notebook.ativo = True
        notebook.save()

        messages.success(request, f"Notebook {notebook.numero} do '{notebook.equipamento.nome}' marcado como consertado.")

        enviar_telegram(
            f"✅ <b>Notebook consertado</b>\n"
            f"Carrinho: {notebook.equipamento.nome}\n"
            f"Número: {notebook.numero}\n"
            f"Atualizado por: {request.user.get_full_name() or request.user.username}"
        )
    except Exception as e:
        messages.error(request, f"Erro: {e}")

    return redirect('notebooks_quebrados')


@login_required
@staff_member_required
def verificar_carrinho(request):
    """
    Permite que um admin escolha um carrinho e digite os números de notebook
    encontrados dentro dele. Compara com a faixa esperada daquele carrinho
    (numero_inicial..numero_final) e aponta:
      - números corretos (estão na faixa do carrinho)
      - faltando (esperados, mas não digitados)
      - fora de lugar (digitados, mas pertencem à faixa de outro carrinho)
      - não reconhecidos (digitados, mas não pertencem a nenhuma faixa cadastrada)
      - duplicados (digitados mais de uma vez)
    """
    equipamentos = Equipamento.objects.all().order_by('nome')
    resultado = None
    carrinho_selecionado = None

    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numeros_texto = request.POST.get('numeros', '')
        carrinho_selecionado = get_object_or_404(Equipamento, id=equipamento_id)

        numeros_informados = [int(n) for n in re.findall(r'\d+', numeros_texto)]

        contagem = {}
        for n in numeros_informados:
            contagem[n] = contagem.get(n, 0) + 1
        duplicados = sorted([n for n, c in contagem.items() if c > 1])

        numeros_unicos = set(numeros_informados)
        faixa_esperada = carrinho_selecionado.faixa_numeros()

        faltando = sorted(faixa_esperada - numeros_unicos)

        outros_equipamentos = list(equipamentos.exclude(id=carrinho_selecionado.id))

        corretos = []
        fora_de_lugar = []
        nao_reconhecidos = []

        for n in sorted(numeros_unicos):
            if n in faixa_esperada:
                corretos.append(n)
                continue
            dono = next((o for o in outros_equipamentos if n in o.faixa_numeros()), None)
            if dono:
                fora_de_lugar.append({'numero': n, 'pertence_a': dono.nome})
            else:
                nao_reconhecidos.append(n)
        resultado = {
            'corretos': corretos,
            'faltando': faltando,
            'fora_de_lugar': fora_de_lugar,
            'nao_reconhecidos': nao_reconhecidos,
            'duplicados': duplicados,
            'total_esperado': len(faixa_esperada),
            'total_informado': len(numeros_unicos),
        }

        if faltando or fora_de_lugar or nao_reconhecidos:
            linhas = [f"🔍 <b>Divergência no carrinho {carrinho_selecionado.nome}</b>"]
            if faltando:
                linhas.append(f"Faltando: {', '.join(map(str, faltando))}")
            if fora_de_lugar:
                itens = ', '.join(f"{f['numero']} (é do {f['pertence_a']})" for f in fora_de_lugar)
                linhas.append(f"Fora de lugar: {itens}")
            if nao_reconhecidos:
                linhas.append(f"Não reconhecidos: {', '.join(map(str, nao_reconhecidos))}")
            enviar_telegram("\n".join(linhas))

    return render(request, 'verificar_carrinho.html', {
        'equipamentos': equipamentos,
        'resultado': resultado,
        'carrinho_selecionado': carrinho_selecionado,
    })

@login_required
@staff_member_required
def atualizar_faixa_numeracao(request):
    """Atualiza numero_inicial/numero_final de um carrinho a partir da página
    de Verificar Carrinhos, e volta para ela (não para o mural)."""
    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numero_inicial = request.POST.get('numero_inicial')
        numero_final = request.POST.get('numero_final')
        try:
            equip = Equipamento.objects.get(id=equipamento_id)
            equip.numero_inicial = int(numero_inicial) if numero_inicial else None
            equip.numero_final = int(numero_final) if numero_final else None
            equip.save()
            messages.success(request, f"Faixa de numeração de '{equip.nome}' atualizada!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('verificar_carrinho')


@login_required
@staff_member_required
def alternar_status_notebook(request):
    if request.method == "POST":
        equipamento_id = request.POST.get('equipamento_id')
        numero = request.POST.get('numero')
        try:
            equip = Equipamento.objects.get(id=equipamento_id)
            numero_int = int(numero)

            notebook, created = Notebook.objects.get_or_create(
                equipamento=equip,
                numero=numero_int,
                defaults={'ativo': False},
            )
            if not created:
                notebook.ativo = not notebook.ativo
                notebook.save()

            status = "ativo" if notebook.ativo else "quebrado"
            messages.success(request, f"Notebook {numero_int} do '{equip.nome}' marcado como {status}.")

            if not notebook.ativo:
                enviar_telegram(
                    f"🔧 <b>Notebook marcado como quebrado</b>\n"
                    f"Carrinho: {equip.nome}\n"
                    f"Número: {numero_int}"
                )
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('verificar_carrinho')

@login_required
@staff_member_required
def painel_reservas_dia(request):
    """Painel de cards das reservas — filtrável por data."""
    data_param = request.GET.get('data')
    try:
        data_sel = datetime.strptime(data_param, '%Y-%m-%d').date() if data_param else date.today()
    except ValueError:
        data_sel = date.today()

    reservas = Reserva.objects.filter(
        data_uso=data_sel
    ).select_related('professor', 'equipamento', 'sala').order_by('horario_inicio')

    return render(request, 'painel_reservas.html', {
        'reservas': reservas,
        'data_atual': data_sel.strftime('%Y-%m-%d'),
        'data_exibicao': data_sel.strftime('%d/%m/%Y'),
    })

@login_required
@staff_member_required
def pendentes_numeracao(request):
    """Painel admin: reservas por quantidade que ainda não tiveram a numeração preenchida."""
    pendentes = Reserva.objects.filter(
        quantidade__isnull=False,
        status__in=['confirmada', 'pendente'],
        numeracao_preenchida=False,
    ).select_related('professor', 'equipamento', 'sala').order_by('data_uso', 'horario_inicio')

    return render(request, 'pendentes_numeracao.html', {
        'pendentes': pendentes,
    })

@login_required
@staff_member_required
def painel_reservas_quantidade(request):
    """Painel de cards com todas as reservas por quantidade (notebooks avulsos) — filtrável por data."""
    data_param = request.GET.get('data')
    try:
        data_sel = datetime.strptime(data_param, '%Y-%m-%d').date() if data_param else date.today()
    except ValueError:
        data_sel = date.today()

    reservas = Reserva.objects.filter(
        data_uso=data_sel,
        quantidade__isnull=False,
    ).select_related(
        'professor', 'equipamento', 'sala'
    ).prefetch_related(
        'numeros_quantidade'
    ).order_by('horario_inicio')

    return render(request, 'painel_reservas_quantidade.html', {
        'reservas': reservas,
        'data_atual': data_sel.strftime('%Y-%m-%d'),
        'data_exibicao': data_sel.strftime('%d/%m/%Y'),
    })

@login_required
@staff_member_required
def menu_ajax(request):
        return render(request, "partials/menu.html")

@login_required
@staff_member_required
def aprovar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    reserva.status = 'confirmada'
    reserva.save()
    messages.success(request, f"Reserva de {reserva.professor.username} aprovada!")

    aprovador = request.user.get_full_name() or request.user.username
    enviar_telegram(
        f"✅ <b>Reserva aprovada</b>\n"
        f"Professor: {reserva.professor.get_full_name() or reserva.professor.username}\n"
        f"Data: {reserva.data_uso.strftime('%d/%m/%Y')}\n"
        f"Equipamento: {reserva.equipamento.nome}\n"
        f"Aprovado por: {aprovador}"
    )
    return redirect('mural')

@login_required
@staff_member_required
def recusar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    reserva.status = 'recusada'
    reserva.save()
    messages.warning(request, f"Reserva de {reserva.professor.username} recusada.")

    recusador = request.user.get_full_name() or request.user.username
    enviar_telegram(
        f"❌ <b>Reserva recusada</b>\n"
        f"Professor: {reserva.professor.get_full_name() or reserva.professor.username}\n"
        f"Data: {reserva.data_uso.strftime('%d/%m/%Y')}\n"
        f"Equipamento: {reserva.equipamento.nome}\n"
        f"Recusado por: {recusador}"
    )
    return redirect('mural')

@staff_member_required
def reserva_fixas_web(request):
    if request.method == 'POST':
        form = ReservaFixaForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            grupo = uuid.uuid4()
            criadas, conflitos = 0, 0
            for dia_semana in [int(x) for x in d['dias_semana']]:
                data_atual = d['data_inicio']
                data_atual += timedelta(days=(dia_semana -data_atual.weekday()) % 7)
                while data_atual <= d['data_fim']:
                    existe = Reserva.objects.filter(
                        equipamento =d ['equipamento'], data_uso = data_atual,
                        horario_inicio =d ['horario_inicio'] 
                    ).exists()
                    if existe:
                        conflitos += 1
                    else:
                        Reserva.objects.create(
                            professor=d['professor'], equipamento=d['equipamento'],
                            sala= d['sala'], data_uso = data_atual,
                            horario_inicio=d['horario_inicio'], horario_fim=d['horario_fim'],
                            status = 'confirmada',
                            grupo_fixo = grupo,
                        )
                        criadas += 1
                    data_atual += timedelta(days=7)
            messages.success(request, f"{criadas} reservas crisadas. {conflitos} conflitos ignorados (já existiam).")
            return redirect('reservas_fixas_web')
    else:
        form = ReservaFixaForm()
        
    # monta os grupos para exibir na tabela
    ids = Reserva.objects.filter(grupo_fixo__isnull=False).values_list('grupo_fixo', flat=True).distinct()
    grupos = []
    for grupo_id in ids:
        qs = Reserva.objects.filter(grupo_fixo=grupo_id).select_related('professor', 'sala', 'equipamento').order_by('data_uso')
        if not qs.exists():
            continue
        primeira, ultima = qs.first(), qs.last()
        grupos.append({
            'id': grupo_id,
            'professor': primeira.professor.get_full_name() or primeira.professor.username,
            'sala': primeira.sala.nome,
            'equipamento': primeira.equipamento.nome,
            'horario_inicio': primeira.horario_inicio.strftime('%H:%M'),
            'periodo': f"{primeira.data_uso.strftime('%d/%m/%Y')} - {ultima.data_uso.strftime('%d/%m/%Y')}",
            'qtd_reservas': qs.count(),
        })

    return render(request, 'reservas_fixas.html', {'form': form, 'grupos': grupos})
@login_required
@staff_member_required
def lista_reservas_fixas(request):
    ids = Reserva.objects.filter(grupo_fixo__isnull=False).values_list('grupo_fixo', flat=True).distinct()
    grupos = []
    for grupo_id in ids:
        qs = Reserva.objects.filter(grupo_fixo=grupo_id).select_related('professor', 'sala', 'equipamento').order_by('data_uso')
        if not qs.exists():
            continue
        primeira, ultima = qs.first(), qs.last()
        grupos.append({
            'id': grupo_id,
            'professor': primeira.professor.get_full_name() or primeira.professor.username,
            'sala': primeira.sala.nome,
            'equipamento': primeira.equipamento.nome,
            'horario':f"{primeira.horario_inicio.strftime('%H:%M')} - {primeira.horario_fim.strftime('%H:%M')}",
            'total': qs.count(),
            'de': primeira.data_uso,
            'ate': ultima.data_uso,
        })
    return render(request, 'reservas_fixas.html', {'grupos': grupos})

@staff_member_required
def excluir_reserva_fixa(request, grupo_id):
    if request.method == 'POST':
        qtd, _ = Reserva.objects.filter(grupo_fixo = grupo_id).delete()
        messages.success(request, f"{qtd} reservas excluidas de todos os dias. ")
    return redirect('reserva_fixas_web')


@login_required
@staff_member_required
def analise_sistema(request):
    """Painel de análise estatística de fichas, professores e equipamentos."""
    from django.contrib.auth.models import User
    from django.db.models import Count
    from django.db.models.functions import ExtractWeekDay

    # 1. Total de reservas e fichas
    total_reservas = Reserva.objects.count()
    reservas_confirmadas = Reserva.objects.filter(status='confirmada').count()
    total_fichas_entregues = RegistroUso.objects.values('reserva').distinct().count()

    # 2. Professores com reservas confirmadas e fichas entregues
    professores_stats = []
    todos_professores = User.objects.filter(reserva__isnull=False).distinct()

    for prof in todos_professores:
        res_total = Reserva.objects.filter(professor=prof, status='confirmada').count()
        fichas_entregues = RegistroUso.objects.filter(reserva__professor=prof).values('reserva').distinct().count()
        taxa = (fichas_entregues / res_total * 100) if res_total > 0 else 0
        professores_stats.append({
            'nome': prof.get_full_name() or prof.username,
            'reservas_confirmadas': res_total,
            'fichas_entregues': fichas_entregues,
            'taxa_entrega': round(taxa, 1)
        })

    professores_mais_entregues = sorted(professores_stats, key=lambda x: x['fichas_entregues'], reverse=True)
    professores_menos_entregues = sorted(professores_stats, key=lambda x: x['fichas_entregues'])

    # 3. Status dos Equipamentos e Notebooks Quebrados
    equipamentos = Equipamento.objects.all()
    equipamentos_data = []
    for eq in equipamentos:
        total_nb = eq.quantidade_ativa() + Notebook.objects.filter(equipamento=eq, ativo=False).count()
        quebrados_nb = Notebook.objects.filter(equipamento=eq, ativo=False).count()
        ativos_nb = eq.quantidade_ativa()
        equipamentos_data.append({
            'nome': eq.nome,
            'tipo': eq.get_tipo_display(),
            'total': total_nb,
            'ativos': ativos_nb,
            'quebrados': quebrados_nb
        })

    # --- Dicionário de nomes de dias e ordem de exibição (Segunda a Domingo) ---
    dias_semana_nomes = {
        1: 'Domingo', 2: 'Segunda-feira', 3: 'Terça-feira', 4: 'Quarta-feira',
        5: 'Quinta-feira', 6: 'Sexta-feira', 7: 'Sábado'
    }
    ordem_dias = [2, 3, 4, 5, 6, 7, 1]  # Segunda -> Domingo

    # 4. Total de reservas por dia da semana (visão geral / gráfico)
    reservas_por_dia_semana_raw = (
        Reserva.objects
        .filter(status='confirmada')
        .annotate(dia_semana=ExtractWeekDay('data_uso'))
        .values('dia_semana')
        .annotate(total=Count('id'))
    )
    totais_por_dia_num = {item['dia_semana']: item['total'] for item in reservas_por_dia_semana_raw}
    reservas_por_dia_semana = [
        {'dia': dias_semana_nomes[num], 'total': totais_por_dia_num.get(num, 0)}
        for num in ordem_dias if totais_por_dia_num.get(num, 0) > 0
    ]
    reservas_por_dia_semana_ord = sorted(reservas_por_dia_semana, key=lambda x: x['total'], reverse=True)
    dia_semana_mais = reservas_por_dia_semana_ord[0] if reservas_por_dia_semana_ord else None
    dia_semana_menos = reservas_por_dia_semana_ord[-1] if reservas_por_dia_semana_ord else None

    # 5. Detalhamento: reservas por Dia da Semana + Horário de Início
    #    (para a coordenação identificar o melhor dia/horário para provas e atividades)
    matriz_bruta = (
        Reserva.objects
        .filter(status='confirmada')
        .annotate(dia_semana=ExtractWeekDay('data_uso'))
        .values('dia_semana', 'horario_inicio')
        .annotate(total=Count('id'))
        .order_by('dia_semana', 'horario_inicio')
    )

    horarios_por_dia_num = {}
    for item in matriz_bruta:
        horarios_por_dia_num.setdefault(item['dia_semana'], []).append({
            'horario': item['horario_inicio'],
            'total': item['total']
        })

    reservas_dia_horario = []
    for num in ordem_dias:
        horarios = horarios_por_dia_num.get(num)
        if not horarios:
            continue
        horarios_ordenados = sorted(horarios, key=lambda x: x['horario'])  # ordem cronológica
        horarios_por_movimento = sorted(horarios, key=lambda x: x['total'])  # menor -> maior
        total_dia = sum(h['total'] for h in horarios)
        reservas_dia_horario.append({
            'dia': dias_semana_nomes[num],
            'total_dia': total_dia,
            'horarios': horarios_ordenados,
            'horario_menos_movimentado': horarios_por_movimento[0],
            'horario_mais_movimentado': horarios_por_movimento[-1],
        })

    # 6. Ranking geral das 10 combinações (dia + horário) com MENOS reservas
    #    -> sugestão direta de melhor momento para agendar prova/atividade
    ranking_menos_movimentado = sorted(
        [
            {
                'dia': dias_semana_nomes[item['dia_semana']],
                'horario': item['horario_inicio'],
                'total': item['total']
            }
            for item in matriz_bruta
        ],
        key=lambda x: x['total']
    )[:10]

    context = {
        'total_reservas': total_reservas,
        'reservas_confirmadas': reservas_confirmadas,
        'total_fichas_entregues': total_fichas_entregues,
        'professores_mais_entregues': professores_mais_entregues[:5],
        'professores_menos_entregues': professores_menos_entregues[:5],
        'equipamentos_data': equipamentos_data,
        'dia_semana_mais': dia_semana_mais,
        'dia_semana_menos': dia_semana_menos,
        'reservas_por_dia_semana': reservas_por_dia_semana,
        'reservas_dia_horario': reservas_dia_horario,
        'ranking_menos_movimentado': ranking_menos_movimentado,
    }
    return render(request, 'analise_sistema.html', context)